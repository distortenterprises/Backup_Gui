#!/usr/bin/env python

# This application is a UAT testing dashboard for managing communications with Testers via sms and email.
# To use this an SMS service is required.
# Import contacts from contacts spreadsheet
from flask import Flask, request, render_template, flash, redirect
from flask.ext.wtf import Form
from wtforms import StringField, SubmitField
from wtforms.validators import Required
from flask_script import Manager
import time
import sys
import openpyxl
from textmagic.rest import TextmagicRestClient
from twilio.rest import TwilioRestClient
import telnetlib
import paramiko
import smtplib
import re
import string
from colours import colour
import os
import threading
import database


print "loading global variables!!"
time.sleep(1)
ftpserver = "localhost"
ftpuser = "benwillett"
ftppass = "Burton!123"
user = "09477258"
pwd1 = "SkiCity3192!"
#pwd2 = "privilege_exec_mode_password"
startup = "startup-config"
running = "running-config"
startUnc = "/Users/benwillett/Desktop/ITWork/Caulfield_Grammar/configs/"
runUnc = "/Users/benwillett/Desktop/ITWork/Caulfield_Grammar/configs/"
mailHost = "outlook.office365.com"
fromEmail = "###@###"
toEmail = "###@###"
emailPass = "#####"
webLogin = "Test"
webPassword ="Password"

print "********************************************************"
print "********************************************************"
print "**        Network backup and config sync tool         **"
print "**  ************************************************  **"
print "**  ************************************************  **"
print "**  ************************************************  **"
print "**          Created by benw@techcamp.com.au           **"
print "********************************************************"
print "********************************************************"

time.sleep(1)

print "Connecting to Mail host " + mailHost + ", please wait!!"

# try:
#     smtpObj = smtplib.SMTP(mailHost, 587)
#     smtpObj.ehlo()
#     smtpObj.starttls()
#     smtpObj.login(fromEmail, emailPass)
#     answer = ()
#     print "Connected to " + mailHost + " Successfully"
# except:
#     print "Unable to connect to " + mailHost

#Use for loop to telnet into each routers and execute commands


app = Flask(__name__)
manager = Manager(app)
app.config['SECRET_KEY'] = 'SkiCity3192!'

y = dict()
users = dict()


y['ip'] = []
y['mobile'] = []
y['email'] = []
y['deviceName'] = []
y['status'] = []
y['messageSent'] = []
y['emailSent'] = []
y['backupFailed'] = []

users['EMAIL'] = []
users['FIRSTNAME'] = []
users['SURNAME'] = []
users['PASSWORD'] = []
users['AGE'] = []
users['GENDER'] = []
users['TOWN'] = []
users['COUNTRY'] = []
users['POSTCODE'] = []
usersFile = "users.xlsx"
userFilePath = "/app/"

try:
    message = "importing users from spreadsheet"
    print (message)
    time.sleep(1)
    usersWb = openpyxl.load_workbook(userFilePath + usersFile)
    print"loaded"
    usersSheet = usersWb.get_sheet_by_name("Sheet 1")
    print"sheet 1"
except:
    message = "could not import users from spreadsheet"
    time.sleep(1)

print(usersSheet)
time.sleep(300)
userRowCount = 0
generator = usersSheet
columnA = generator.columns[0]
columnB = generator.columns[1]
columnC = generator.columns[2]
columnD = generator.columns[3]
columnE = generator.columns[4]
columnF = generator.columns[5]
columnG = generator.columns[6]
columnH = generator.columns[7]
columnI = generator.columns[8]


for i in range(0, usersSheet.max_row):
    if (str(columnA[userRowCount].value)) == "None":
        cell = False
    else:
        cell = True
    if cell == True:
        firstnames = (str(columnA[userRowCount].value))
        users['EMAIL'].append(firstnames)
        print(users['EMAIL'])

        firstName = (str(columnB[userRowCount].value))
        users['FIRSTNAME'].append(firstName)
        print(users['FIRSTNAME'])

        surName = (str(columnC[userRowCount].value))
        users['SURNAME'].append(surName)
        print(users['SURNAME'])

        passwords = (str(columnD[userRowCount].value))
        users['PASSWORD'].append(passwords)
        print(users['PASSWORD'])

        ages = (str(columnE[userRowCount].value))
        users['AGE'].append(ages)
        print(users['AGE'])

        gender = (str(columnF[userRowCount].value))
        users['GENDER'].append(gender)
        print(users['GENDER'])

        towns = (str(columnG[userRowCount].value))
        users['TOWN'].append(towns)
        print(users['TOWN'])

        country = (str(columnH[userRowCount].value))
        users['COUNTRY'].append(country)
        print(users['COUNTRY'])

        postcode = (str(columnI[userRowCount].value))
        users['POSTCODE'].append(postcode)
        print(users['POSTCODE'])

    else:
        pass
    userRowCount += 1

username = "XXXXX"
token = "XXXXX"
client = TextmagicRestClient(username, token)

session = False
credentialsP = "null"
credentialsU = "null"
error = ""

exitFlag = 0

class myThread (threading.Thread):
    def __init__(self, threadID, name, counter):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.counter = counter
    def run(self):
        print "Starting " + self.name
        for ip in ipList:
            getconfig(ip)
        print "Exiting " + self.name


# class NameForm(Form):
#     name = StringField('What is your name?', validators=[Datarequired()])
#     submit = SubmitField('Submit')


@app.route('/' , methods=['GET','POST'])
def default():
    global error
    error = ""
    return render_template('login.html', ERROR=error)


@app.route('/login' , methods=['POST'])
def login():
    global error
    global session
    global credentialsP
    global credentialsU
    global users
    credentialsU = request.form['USERNAME']
    credentialsP = request.form['PASSWORD']
    print(credentialsU)
    print(credentialsP)

    count = 0
    while session == False:
        for v in users["EMAIL"]:
            print(v)
            print(count)

            if v == credentialsU and credentialsP == users["PASSWORD"][count]:
                error = ""
                session = True
                return render_template('uat.html', Y=y, SESSION=session, USERNAME=credentialsU, ERROR=error)

            else:
                pass

            count += 1

        error = "Try again"
        session = False
        return render_template('login.html', Y=y, SESSION=session, USERNAME=credentialsU, ERROR=error)



@app.route('/logout' , methods=['GET','POST'])
def logout():
    global session
    session = False
    return render_template('login.html')


@app.route('/uat' , methods=['GET','POST'])
def uat():
    global y
    global session
    global credentialsP
    global credentialsU
    if session == True:
        return render_template('uat.html', Y=y, SESSION=session, USERNAME=credentialsU)
    else:
        return redirect('/')


@app.route('/admin', methods=['GET', 'POST'])
def admin():
        global users
        global session
        global credentialsU
        print("admin")
        if session == True:
            return render_template('admin.html', USERS=users, USERNAME=credentialsU)
        else:
            return redirect('/')


@app.route('/adduser', methods=['GET', 'POST'])
def addUser():
        return render_template('adduser.html')


@app.route('/deleteadmin', methods=['GET', 'POST'])
def deleteadmin():
    global users
    global credentialsU
    global session
    deleteA = request.form['DELETE']
    print(deleteA + " in deleteA variable")
    x = False
    count = 0
    while x == False:
        print("at start of while loop")
        for each in users['EMAIL']:
            print(each)
            if each == deleteA:
                users['EMAIL'].remove(deleteA)
                print("tried to delete")
                x = True
            else:
                pass
        count += 1
    print("admin")
    if session == True:
        return redirect('/admin')
    else:
        return redirect('/')


@app.route('/submituser', methods=['GET', 'POST'])
def submituser():
    global users
    global error
    error = ""
    email = request.form['EMAIL']
    firstName = request.form['FIRSTNAME']
    surname = request.form['SURNAME']
    password = request.form['PASSWORD']
    age = request.form['AGE']
    gender = request.form['GENDER']
    town = request.form['TOWN']
    country = request.form['COUNTRY']
    postCode = request.form['POSTCODE']

    users['EMAIL'].append(email)
    users['FIRSTNAME'].append(firstName)
    users['SURNAME'].append(surname)
    users['PASSWORD'].append(password)
    users['AGE'].append(age)
    users['GENDER'].append(gender)
    users['TOWN'].append(town)
    users['COUNTRY'].append(country)
    users['POSTCODE'].append(postCode)

    print(gender,country,postcode)
    return render_template('login.html', ERROR=error)


@app.route('/fileimport' , methods=['POST'])
def fileImport():

    fileImport = request.form['IMPORT']

    spreadsheet = fileImport
    spreadsheetPath = "/Users/benwillett/PycharmProjects/Backup_Gui/"
    print(spreadsheet)

    print(fileImport)
    print("file import ok")

    try:
        y['ip'] = []
        y['mobile'] = []
        y['email'] = []
        y['deviceName'] = []
        y['status'] = []
        y['messageSent'] = []
        y['emailSent'] = []
        y['backupFailed'] = []

        message = "importing devices from spreadsheet"
        print (message)
        time.sleep(1)
        wb = openpyxl.load_workbook(spreadsheetPath + spreadsheet)
        print"loaded"
        sheet = wb.get_sheet_by_name("Sheet1")
        print"sheet 1"
    except:
        message = "could not import devices from spreadsheet"
        time.sleep(1)
        sys.exit('operation completed')

    rowCount = 0
    columnA = sheet.columns[0]
    columnB = sheet.columns[1]
    columnC = sheet.columns[2]
    columnD = sheet.columns[3]

    for i in range(0, sheet.max_row):
        if (str(columnA[rowCount].value)) == "None":
            cell = False
        else:
            cell = True
        if cell == True:
            deviceIp = (str(columnA[rowCount].value))
            y['ip'].append(deviceIp)
            print(y['ip'])

            appResponse = (str(columnB[rowCount].value))
            y['deviceName'].append(appResponse)
            print(y['deviceName'])

            emailAddress = (str(columnC[rowCount].value))
            y['email'].append(emailAddress)
            print(y['email'])

            mobileNum = (str(columnD[rowCount].value))
            y['mobile'].append(mobileNum)
            print(y['mobile'])


            print(y['status'])
        else:
            pass
        rowCount += 1

    print("redirecting")
    return redirect('/uat')


@app.route('/message' , methods=['POST'])
def message():

    global myThread
    global y
    global ipList, emailList

    print(y["status"])
    sms = request.form['submit']

    checkedSms = request.form.getlist('ENABLE')
    print(checkedSms)
    ipList = list(set(checkedSms))
    print("ok 1")
    print(ipList)

    try:
        ipList.remove("None")
    except:
        pass

    checkedEmail = request.form.getlist('EMAIL')
    emailList = list(set(checkedEmail))
    print("ok list 2")

    try:
        emailList.remove("None")
    except:
        pass

    print("getconfig function run")



    thread1 = myThread(1,"Thread-1", 1)
    # Start new Threads
    thread1.start()

    print("redirecting")
    return redirect('/uat')


def sshconnect(ip):
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(ip, username=user, password=pwd1)
    # inventory[ipAddresses]['name'] = "empty"
    remote_conn = client.invoke_shell()
    print("Interactive SSH session established")
     # Strip the initial router prompt
    output = remote_conn.recv(1000)
    # See what we have
    print output
    # Turn off paging
    #disable_paging(remote_conn)
    # Now let's try to send the router a command
    remote_conn.send("\n")
    # Wait for the command to complete
    time.sleep(2)
    output = remote_conn.recv(1100)
    print output


def getconfig(ip):
    global y
    global ipList, emailList, startup, running, ftpuser, ftppass, ftpserver, startUnc, runUnc


    try:
        print("Backing up device " + ip)
        cmd1 = "en"
        cmd2 = "copy " + startup + " ftp://" + ftpuser + ":" + ftppass + "@" + ftpserver + startUnc + ip + ".txt"
        cmd3 = "copy " + running + " ftp://" + ftpuser + ":" + ftppass + "@" + ftpserver + runUnc + ip + ".txt"

        try:
            telnet = telnetlib.Telnet(ip, 10)
            try:
                time.sleep(1)
                telnet.write(user.encode('ascii') + b"\n")
                time.sleep(1)
                telnet.write(pwd1.encode('ascii') + b"\n")
                time.sleep(1)
                telnet.write(cmd1.encode('ascii') + b"\n")
                time.sleep(1)
                telnet.write(cmd2.encode('ascii') + b"\n")
                time.sleep(2)
                telnet.write(b"\n")
                time.sleep(1)
                telnet.write(b"\n")
                time.sleep(1)
                telnet.write(cmd3.encode('ascii') + b"\n")
                time.sleep(2)
                telnet.write(b"\n")
                time.sleep(1)
                telnet.write(b"\n")
                time.sleep(1)
                telnet.close()
                time.sleep(1)
                print "Successfully transfered config for " + ip
                y['messageSent'].append(ip)
            except:
                print "Unable to authenticate to " + ip
                y['backupFailed'].append(ip)
                pass
        except:
            print "Unable to telnet into " + ip
            try:
                print "Attempting to SSH into " + ip
                sshconnect(ip)
            except:
                print "Unable to SSH into " + ip
                y['backupFailed'].append(ip)
                pass

    except:
        pass

    # for each in emailList:
    #     try:
    #         print("Sending EMAIL to " + each + " MESSAGE:" + sms)
    #         print("EMAIL sent successfully to" + each)
    #         y['emailSent'].append(each)
    #
    #     except:
    #         print("could not connect to SMTP Server, EMAIL not sent to " + each)
    #         pass
    #
    #
    # try:
    #     print("got this far")
    #     status = list(set(checkedStatus))
    #     k = y['status'] + status
    #     y['status'] = k
    #     print(k)
    #
    # except:
    #     print("problem adding status")
    # print(y['status'])



if __name__ == '__main__':
    manager.run()
    app.run(debug=True)
